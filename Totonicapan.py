import pdfplumber
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import streamlit as st
import re
import io

# --- HELPER FUNCTIONS ---
def normalize_text(text):
    if not text: return ""
    nfd = unicodedata.normalize('NFD', str(text))
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn').lower()

def squish_text(text):
    """Aggressively removes ALL spaces, punctuation, hyphens, and hidden characters for a 100% reliable match."""
    if not text: return ""
    t = normalize_text(text)
    return re.sub(r'[^a-z0-9]', '', t)

def safe_float(val):
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-': return 0.0
    s = s.replace(',', '') 
    s = re.sub(r'[^\d\.\-]', '', s) 
    if s.count('.') > 1:
        parts = s.rsplit('.', 1)
        s = parts[0].replace('.', '') + '.' + parts[1]
    try: return float(s)
    except ValueError: return 0.0

def clean_currency(value):
    if not value: return 0.0
    raw = str(value).strip().replace(' ', '')
    raw = re.sub(r'[^\d\.,]', '', raw)
    if not raw: return 0.0
    
    if re.search(r',\d{1,2}$', raw):
        parts = raw.rsplit(',', 1)
        raw = parts[0].replace('.', '').replace(',', '') + '.' + parts[1]
    else:
        raw = raw.replace(',', '')
        
    if raw.count('.') > 1:
        parts = raw.rsplit('.', 1)
        raw = parts[0].replace('.', '') + '.' + parts[1]
        
    try: return float(raw)
    except ValueError: return 0.0

def extract_value_from_row(row_list, total_idx):
    if total_idx != -1 and len(row_list) > total_idx:
        val = clean_currency(row_list[total_idx])
        if val > 0: return val
    for item in reversed(row_list):
        val = clean_currency(item)
        if val > 0: return val
    return 0.0

def get_master_cell(ws, r_idx, c_idx):
    cell = ws.cell(row=r_idx, column=c_idx)
    if type(cell).__name__ == 'MergedCell':
        for m_range in ws.merged_cells.ranges:
            if cell.coordinate in m_range:
                return ws.cell(row=m_range.min_row, column=m_range.min_col)
    return cell

def find_description_in_row(row):
    """
    Find the product description in a row by finding the longest text cell
    that contains letters (not pure numbers/symbols).
    """
    best_candidate = ""
    best_score = 0
    
    for cell in row:
        if cell is None:
            continue
        
        cell_str = str(cell).strip()
        if not cell_str:
            continue
        
        # Skip obvious non-descriptions
        cell_upper = cell_str.upper()
        if cell_upper in ['BIEN', 'SERVICIO', 'B/S']:
            continue
        if cell_upper.startswith('IVA ') or cell_upper.startswith('ISR '):
            continue
        
        # Check if it's a pure number
        try:
            float(cell_str.replace(',', '.').replace(' ', ''))
            # It's a number - only use if nothing else found
            continue
        except ValueError:
            # Not a pure number - this is good!
            pass
        
        # Score this cell based on how likely it is to be a description
        # Longer text with more letters = higher score
        letter_count = sum(1 for c in cell_str if c.isalpha())
        
        # Must have at least some letters to be a description
        if letter_count < 3:
            continue
        
        # Score = number of letters (prefer text over numbers)
        score = letter_count
        
        if score > best_score:
            best_score = score
            best_candidate = cell_str
    
    return best_candidate

def merge_split_rows(tables):
    """
    Merges rows that were split due to white lines in PDF tables.
    
    Detects continuation rows (rows with only description text but no item number/value)
    and merges them back into the previous data row's description.
    
    Example:
        Row N:   ['23', None, 'Bien', '32', 'UNIDADES DE', '5.50', ..., '176.00']
        Row N+1: [None, '', '', '', 'AGUACATE', '', '', '', '', '']  ← continuation
        
        After merge:
        Row N:   ['23', None, 'Bien', '32', 'UNIDADES DE AGUACATE', '5.50', ..., '176.00']
        Row N+1: removed
    """
    if not tables:
        return tables
    
    merged = []
    i = 0
    while i < len(tables):
        current_row = list(tables[i]) if tables[i] else []
        
        # Look ahead to merge any continuation rows
        j = i + 1
        while j < len(tables):
            next_row = tables[j]
            if not next_row:
                break
            
            # Check if next_row is a continuation row:
            # 1. No item number (no digit) in first 5 cells
            # 2. No numeric values anywhere
            # 3. Has at least some text content
            has_item_number = False
            for cell in next_row[:5]:
                if cell:
                    cell_str = str(cell).strip()
                    if cell_str.isdigit():
                        has_item_number = True
                        break
            
            has_numeric_value = False
            text_fragments = []
            for cell in next_row:
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                if not cell_str:
                    continue
                # Check if it's a number
                try:
                    val = float(cell_str.replace(',', '.').replace(' ', ''))
                    if val > 0:
                        has_numeric_value = True
                        break
                except ValueError:
                    # Not a number - could be description text
                    cell_upper = cell_str.upper()
                    if len(cell_str) >= 3 and cell_upper not in ['BIEN', 'SERVICIO', 'B/S']:
                        if not cell_upper.startswith('IVA') and not cell_upper.startswith('ISR'):
                            text_fragments.append(cell_str)
            
            # If it's a continuation row, merge text into current row's description
            if not has_item_number and not has_numeric_value and text_fragments:
                continuation_text = " ".join(text_fragments)
                
                # Find description cell in current row and append the continuation
                for k, cell in enumerate(current_row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    if not cell_str:
                        continue
                    # Skip non-description cells
                    cell_upper = cell_str.upper()
                    if cell_upper in ['BIEN', 'SERVICIO', 'B/S']:
                        continue
                    if cell_upper.startswith('IVA') or cell_upper.startswith('ISR'):
                        continue
                    # Skip pure numbers
                    try:
                        float(cell_str.replace(',', '.').replace(' ', ''))
                        continue
                    except ValueError:
                        pass
                    # This is the description cell - append continuation
                    if len(cell_str) >= 3:
                        current_row[k] = cell_str + " " + continuation_text
                        break
                
                j += 1  # Move to next row, continue checking for more continuations
            else:
                # Not a continuation, stop merging
                break
        
        merged.append(current_row)
        i = j  # Skip past any merged continuation rows
    
    return merged

def fuzzy_match_category(description, cultivados, abarrotes, threshold=80):
    """
    Accent-insensitive, word-boundary matching with Spanish plural tolerance.
    Handles pdfplumber multi-line cells and concatenated units like 'espagueti180g'.
    """
    if not description:
        return ('unmatched', None)

    desc_norm = normalize_text(description)
    # Collapse newlines/tabs/repeated whitespace from multi-line PDF cells
    desc_norm = re.sub(r'\s+', ' ', desc_norm).strip()
    # Split letter/digit runs so "espagueti180g" -> "espagueti 180 g"
    desc_norm = re.sub(r'([a-z])(\d)', r'\1 \2', desc_norm)
    desc_norm = re.sub(r'(\d)([a-z])', r'\1 \2', desc_norm)

    # (e?s)? tolerates Spanish plurals: banano/bananos, limon/limones, etc.
    for kw in cultivados:
        if re.search(r'\b' + re.escape(kw) + r'(e?s)?\b', desc_norm):
            return ('agricultura', kw)
    for kw in abarrotes:
        if re.search(r'\b' + re.escape(kw) + r'(e?s)?\b', desc_norm):
            return ('abarrotes', kw)

    return ('unmatched', None)

# --- TRUCO CSS PARA TRADUCIR LA INTERFAZ A ESPAÑOL ---
st.markdown("""
    <style> 
        div[data-testid="stFileUploader"] label p {
            font-size: 40px !important;
        }
    </style>
""", unsafe_allow_html=True)

# --- WEB UI ---
st.title("🇬🇹 MAGA: Procesador de Facturas por la LAE: Totonicapán")

# Municipality selector - user must specify which municipality the receipts belong to
MUNICIPIOS_OPCIONES = {
    "Totonicapán": 1,
    "San Cristóbal Totonicapán": 2,
    "San Francisco El Alto": 3,
    "San Andrés Xecul": 4,
    "Momostenango": 5,
    "Santa María Chiquimula": 6,
    "Santa Lucía La Reforma": 7,
    "San Bartolo Aguas Calientes": 8
}

selected_municipio = st.selectbox(
    label='1. Seleccione el Municipio de las facturas',
    options=["-- Seleccionar municipio --"] + list(MUNICIPIOS_OPCIONES.keys()),
    help="Todas las facturas que suba deben corresponder a este municipio"
)

uploaded_pdfs = st.file_uploader(label='2. Seleccione sus Facturas (PDFs)', type='pdf', accept_multiple_files=True)
uploaded_xlsx = st.file_uploader(label='3. Seleccione su Archivo de Excel', type='xlsx')

# Show warning if municipality not selected
municipio_valido = selected_municipio != "-- Seleccionar municipio --"

# Show informational message about municipality selection
if municipio_valido:
    st.info(f"📍 Municipio seleccionado: **{selected_municipio}**. Asegúrese de que todas las facturas correspondan a este municipio.")
else:
    st.warning("⚠️ Por favor seleccione un municipio antes de iniciar el proceso.")

if st.button("INICIAR PROCESO") and uploaded_pdfs and uploaded_xlsx and municipio_valido:
    try:
        # Get municipality info from user selection
        user_m_id = MUNICIPIOS_OPCIONES[selected_municipio]
        user_m_name = selected_municipio
        
        input_buffer = io.BytesIO(uploaded_xlsx.read())
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active 
        
        if "Extra Detalles" not in wb.sheetnames:
            ws_det = wb.create_sheet("Extra Detalles")
            ws_det.append(['Archivo PDF', 'Nombre Emisor', 'NIT Emisor', 'NIT Receptor', 'Num. DTE', 'Municipio', 'Alerta % Abarrotes'])
        else:
            ws_det = wb["Extra Detalles"]
        
        # Create sheet for unmatched items
        if "Items Sin Clasificar" not in wb.sheetnames:
            ws_unmatched = wb.create_sheet("Items Sin Clasificar")
            ws_unmatched.append(['Descripción', 'Municipio', 'Total (Q)', 'Num. DTE'])
        else:
            ws_unmatched = wb["Items Sin Clasificar"]

        # 1. Map Excel Columns dynamically
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=15): 
            for cell in row:
                if type(cell).__name__ == 'MergedCell': continue
                if not cell.value: continue
                val = normalize_text(str(cell.value))
                
                if 'abarrotes' in val: col_map['abar'] = cell.column
                if 'agricultura' in val: col_map['agri'] = cell.column
                if 'escuela' in val or 'establecimiento' in val: col_map['escuelas'] = cell.column
                if 'proveedor' in val or 'productor' in val:
                    base_col, base_row, found_total = cell.column, cell.row, False
                    for r_offset in range(1, 4):
                        for c_offset in range(3):
                            sub_cell = ws.cell(row=base_row + r_offset, column=base_col + c_offset)
                            if sub_cell.value and 'total' in normalize_text(str(sub_cell.value)):
                                col_map['productores'] = sub_cell.column
                                found_total = True
                                break
                        if found_total: break
                    if 'productores' not in col_map: col_map['productores'] = base_col

        if 'abar' not in col_map or 'agri' not in col_map:
            st.error(f"No encontré las columnas base en el Excel.")
            st.stop()

        department_name = 'totonicapan'
        # 2. MASTER MUNICIPALITY DICTIONARY
        MUNICIPIOS = {
            1: {"nombre_oficial": "Totonicapán", "alias_pdf": ["totonicapan totonicapan", "totonicapan, totonicapan", "totonicapan"]},
            2: {"nombre_oficial": "San Cristóbal Totonicapán", "alias_pdf": ["san cristobal totonicapan", "san cristobal"]},
            3: {"nombre_oficial": "San Francisco El Alto", "alias_pdf": ["san francisco el alto", "san francisco"]},
            4: {"nombre_oficial": "San Andrés Xecul", "alias_pdf": ["san andres xecul", "san andres"]},
            5: {"nombre_oficial": "Momostenango", "alias_pdf": ["momostenango"]},
            6: {"nombre_oficial": "Santa María Chiquimula", "alias_pdf": ["santa maria chiquimula", "sta maria chiquimula", "santa maria", "sta maria"]},
            7: {"nombre_oficial": "Santa Lucía La Reforma", "alias_pdf": ["santa lucia la reforma", "sta lucia la reforma", "santa lucia", "sta lucia"]},
            8: {"nombre_oficial": "San Bartolo Aguas Calientes", "alias_pdf": ["san bartolo aguas calientes", "san bartolo"]}
        }
        
        search_list = []
        for m_id, data in MUNICIPIOS.items():
            for alias in data["alias_pdf"]:
                search_list.append((alias, m_id, data["nombre_oficial"]))
                
        # CORE FIX: Sorts the list so Totonicapán (ID 1) is ALWAYS evaluated last.
        # Within the other municipalities, sorts by length to catch specific names first.
                search_list.sort(key=lambda x: (
            squish_text(x[2]) == squish_text(department_name),
            -len(x[0])
        ))
        
        EXCEL_MAPPINGS = {
            1: "totonicapán", 2: "san cristobal", 3: "san francisco", 4: "san andres",
            5: "momostenango", 6: "santa maria", 7: "santa lucia", 8: "san bartolo"
        }

        # 3. Map Excel Rows to Municipalities
        row_map = {}
        for row_ex in ws.iter_rows(min_row=5, max_row=150):
            row_text = " ".join([str(c.value) for c in row_ex if c.value and type(c).__name__ != 'MergedCell'])
            row_squished = squish_text(row_text)
            for m_id, search_key in EXCEL_MAPPINGS.items():
                if m_id in row_map: continue
                key_squished = squish_text(search_key)
                if key_squished in row_squished:
                    row_map[m_id] = row_ex[0].row

        batch_totals = {m_id: {'abar': 0.0, 'agri': 0.0, 'emisores': set(), 'receptores': set()} for m_id in MUNICIPIOS.keys()}
        new_count = 0
        skipped_non_standard = []  # Track non-standard receipts
        progress_bar = st.progress(0)

        # 4. Process each PDF
        for i, pdf_file in enumerate(uploaded_pdfs):
            with pdfplumber.open(pdf_file) as pdf:
                text = "".join([p.extract_text() or "" for p in pdf.pages])
                tables = []
                for p in pdf.pages:
                    t = p.extract_table()
                    if t: tables.extend(t)
                
                # Merge rows that were split by white lines in the PDF
                tables = merge_split_rows(tables)

                # VALIDATION: Check if this is a standard SAT factura
                # Standard facturas have specific markers that proformas/cotizaciones don't
                is_standard_factura = False
                
                # Check 1: Must have "Número de DTE" (unique to SAT facturas)
                has_dte = bool(re.search(r'N[úu]mero\s*de\s*DTE', text, re.IGNORECASE))
                
                # Check 2: Must have "NÚMERO DE AUTORIZACIÓN" (SAT authorization)
                has_autorizacion = bool(re.search(r'N[úu]mero\s*de\s*Autorizaci[óo]n', text, re.IGNORECASE))
                
                # Check 3: Must have "Nit Emisor" in standard format
                has_nit_emisor = bool(re.search(r'Nit\s*Emisor', text, re.IGNORECASE))
                
                # Must have at least 2 of the 3 markers to be considered a valid factura
                marker_count = sum([has_dte, has_autorizacion, has_nit_emisor])
                is_standard_factura = marker_count >= 2
                
                if not is_standard_factura:
                    skipped_non_standard.append(pdf_file.name)
                    progress_bar.progress((i + 1) / len(uploaded_pdfs))
                    continue

                dte_m = re.search(r'N[úu]mero\s*de\s*DTE:\s*(\d+)', text, re.IGNORECASE)
                dte_val = dte_m.group(1) if dte_m else pdf_file.name

                # Use the municipality selected by the user (not detected from receipt)
                m_id = user_m_id
                m_name = user_m_name

                if m_id:
                    abar_sum, agri_sum = 0, 0
                    
                    cultivados = [
                        # frutas
                        'banano', 'bananano',                         # triple-n typo
                        'platano', 'pina', 'papaya', 'sandia', 'melon', 'mango',
                        'naranja', 'limon', 'limom', 'limo',          # limon typos
                        'manzana', 'aguacate', 'jamaica', 'tamarindo',
                        'guayaba', 'fresa', 'mora', 'arandano', 'orandano',
                        # verduras / hortalizas
                        'tomate', 'miltomate', 'cebolla', 'zanahoria', 'ejote',
                        'guisquil', 'gusiquil', 'guisqul',            # guisquil typos
                        'guicoy', 'ayote', 'calabaza', 'remolacha', 'repollo',
                        'brocoli', 'brocoly',                          # brocoli typo
                        'coliflor', 'papa', 'camote', 'yuca', 'malanga',
                        'espinaca', 'bledo', 'rabano', 'lechuga', 'pepino',
                        'chipolin', 'chipilin',
                        # hierbas / aromaticas
                        'perejil', 'ajo', 'apio', 'cilantro', 'chipilin', 'oregano', 'romero',
                        'hierba', 'hierba buena', 'hierbabuena', 'hirbabuena',
                        'mashan', 'apazote', 'apasote',                # apazote misspelling
                        'zacate', 'tusa', 'laurel', 'tomio', 'tomillo', 'albahaca',
                        # granos frescos
                        'maiz', 'cebada', 'cabada',                    # cebada typo
                        'trigo', 'arveja', 'haba', 'azote',
                        # chiles cultivados (qualified only — bare "chile" stays unmatched)
                        'chile pimiento', 'chile pimento',             # pimento typo
                        'chile pasa', 'chila pasa',                    # chila typo
                        'chile guaque', 'chile guaca',                 # guaca typo (very common)
                        'chile cobanero', 'chile verde', 'chile jalapeno', 'chile chiltepe',
                        'chile dulce', 'chile morron', 'chile chocolate', 'chile negro', 'achiote', 'chile',
                        # frijol cultivado
                        'frijol ejotero', 'frijol tierno', 'frijol negro', 'frijol vaina real',
                        
                    ]
                    
                    abarrotes = [
                        # semillas secas / procesadas
                        'ajonjoli', 'ajonjolin',                       # ajonjoli variant spelling
                        'pepita', 'pepitoria', 'pepitorio', 'frijol sellado',
                        'mani', 'mania',                              # mani typo
                        # proteina animal
                        'huevo', 'pollo', 'pechuga', 'pierna', 'muslo', 'res', 'carne',
                        'pescado', 'embutido', 'chorizo', 'salchicha', 'jamon',
                        # lacteos
                        'crema', 'leche', 'queso', 'yogur', 'mantequilla', 'margarina',
                        # panaderia
                        'pan', 'pirujo', 'cevada',                              # "pirujo" sometimes appears without "pan"
                        'tostada', 'tortilla', 'galleta', 'chocolate',
                        # pasta / cereales procesados
                        'pasta', 'espagueti', 'fideo', 'macarron',
                        'codito',                                      # catches "pasta codito" / "pasto codito"
                        'avena', 'abena',                              # avena typo
                        'corazon de trigo',
                        'chaomein', 'chow mein', 'chao mein', 'chaumein', 'cahomein',
                        'mosh',                                        # Guatemalan oatmeal (mosh quaker)
                        # harinas / mezclas
                        'maseca', 'incaparina', 'protemas', 'atol', 'harina', 'pinol',
                        # aceites / condimentos
                        'aceite', 'sal', 'azucar', 'vinagre',
                        'achiote', 'achote',                           # achiote typo
                        'canela',
                        'laurel', 'laure',                             # laurel typo
                        'tomillo', 'clavo', 'pimienta', 'comino',
                        'pimiento en polvo',                           # paprika-like: processed, not fresh pimiento
                        # otros
                        'arroz', 'consome', 'concentrado', 'levadura', 'agua pura', 'bebida',
                        # chiles procesados / secos
                        'chile seco', 'chile rojo', 'chile en polvo', 'chile molido',
                        # frijol procesado / seco
                        'frijol negro', 'frijol rojo', 'frijol colorado',  # colorado = rojo variant
                        'frijol blanco', 'frijol en grano', 'frijol seco',
                        #etc.
                        'crayones', 'crayones de madera', 'sacapuntas', 'borradores',
                        'frascos de goma', 'lapiz', 'lapicero', 'lapiceros', 'lapices',
                        'cuadernos', 'espaqueti'
                    ]
                    
                    # Find the Total column and Description column indices
                    # ONLY search in the first table's header rows (first 5 rows max)
                    total_col_idx = -1
                    desc_col_idx = -1
                    
                    if tables:
                        header_rows = tables[:min(5, len(tables))]
                        for row_tbl in header_rows:
                            if not row_tbl: continue
                            for idx, cell in enumerate(row_tbl):
                                if not cell: continue
                                cell_norm = normalize_text(str(cell))
                                
                                # Find Total column (has "Total" and "(Q)")
                                if 'total' in cell_norm and 'descuento' not in cell_norm and '(q)' in cell_norm:
                                    total_col_idx = idx
                                
                                # Find Description column
                                if 'descripcion' in cell_norm:
                                    desc_col_idx = idx
                            
                            if total_col_idx != -1 and desc_col_idx != -1:
                                break
                    
                    # If we didn't find the description column, assume it's index 3
                    if desc_col_idx == -1:
                        desc_col_idx = 3

                    # Process each row in the tables
                    for row_tbl in tables:
                        if not row_tbl: continue
                        
                        # Build full row text for matching
                        row_text = " ".join([str(x) for x in row_tbl if x])
                        row_text_normalized = normalize_text(row_text)
                        
                        # FILTER 1: Skip rows with administrative keywords
                        skip_keywords = ['totales', 'superintendencia', 'datos del certificador', 
                                        'contribuyendo', 'sujeto a pagos', 'no genera derecho',
                                        'descripcion', 'cantidad', 'unitario', 'descuentos', 'impuestos']
                        if any(keyword in row_text_normalized for keyword in skip_keywords):
                            continue
                        
                        # FILTER 2: Check if this looks like a data row
                        # Look for a digit in the first few cells (item numbers like 1, 2, 3...)
                        is_data_row = False
                        for cell in row_tbl[:5]:  # Check first 5 cells (more lenient)
                            if cell:
                                cell_str = str(cell).strip()
                                if cell_str.isdigit():
                                    is_data_row = True
                                    break
                        
                        if not is_data_row:
                            continue
                        
                        # FILTER 3: Skip "artifact rows" caused by white lines splitting a row
                        # These rows have very few non-empty cells (usually just the item number)
                        non_empty_cells = sum(1 for c in row_tbl if c and str(c).strip())
                        if non_empty_cells < 3:
                            continue
                        
                        # Extract the value
                        val = extract_value_from_row(row_tbl, total_col_idx)
                        
                        # Skip rows with zero or invalid value
                        if val <= 0:
                            continue
                        
                        # Use intelligent description finder
                        description = find_description_in_row(row_tbl)
                        
                        # Fallback chain: try progressively more aggressive methods
                        if not description:
                            # Method 1: Try the detected description column
                            if desc_col_idx != -1 and desc_col_idx < len(row_tbl):
                                cell = row_tbl[desc_col_idx]
                                if cell:
                                    description = str(cell).strip()
                        
                        if not description:
                            # Method 2: Try index 3 (standard description column)
                            if len(row_tbl) > 3 and row_tbl[3]:
                                description = str(row_tbl[3]).strip()
                        
                        if not description:
                            # Method 3: Find the longest non-numeric cell
                            longest = ""
                            for cell in row_tbl:
                                if not cell:
                                    continue
                                cell_str = str(cell).strip()
                                # Skip pure numbers
                                try:
                                    float(cell_str.replace(',', '.'))
                                    continue
                                except ValueError:
                                    # Skip very short cells and keywords
                                    if len(cell_str) > len(longest) and cell_str.upper() not in ['BIEN', 'SERVICIO']:
                                        longest = cell_str
                            if longest:
                                description = longest
                        
                        if not description:
                            # Method 4: Just use the longest cell period (even if it's a number)
                            longest = max((str(c).strip() for c in row_tbl if c), key=len, default="")
                            if longest:
                                description = longest
                        
                        if not description:
                            # Method 5: Last resort - use row text
                            description = "REVISAR: " + row_text[:50]
                        
                        # Use fuzzy matching to categorize (using full row text for matching)
                        category, matched_word = fuzzy_match_category(row_text, cultivados, abarrotes, threshold=80)
                        
                        if category == 'agricultura':
                            agri_sum += val
                        elif category == 'abarrotes':
                            abar_sum += val
                        elif category == 'unmatched':
                            # Add ONLY the description to unmatched items sheet
                            ws_unmatched.append([description, m_name, val, dte_val])
                    
                    nit_e_match = re.search(r'Emisor:\s*([0-9Kk\-]+)', text, re.I)
                    nit_r_match = re.search(r'Receptor:\s*([0-9Kk\-]+)', text, re.I)
                    name_e_match = re.search(r'(?:Factura(?:\s*Pequeño\s*Contribuyente)?)\s*\n+(.*?)\n+Nit\s*Emisor', text, re.IGNORECASE | re.DOTALL)
                    
                    nit_e = nit_e_match.group(1).strip() if nit_e_match else "N/A"
                    nit_r = nit_r_match.group(1).strip() if nit_r_match else "N/A"
                    raw_name = re.sub(r'\s+', ' ', name_e_match.group(1).strip() if name_e_match else "N/A")
                    name_e = re.split(r'(?i)n[úu]mero\s*de\s*autorizaci[óo]n', raw_name)[0]
                    name_e = re.split(r'(?i)\bserie\b', name_e)[0].strip()

                    batch_totals[m_id]['abar'] += abar_sum
                    batch_totals[m_id]['agri'] += agri_sum
                    if nit_e != "N/A": batch_totals[m_id]['emisores'].add(nit_e)
                    if nit_r != "N/A": batch_totals[m_id]['receptores'].add(nit_r)

                    total_rec = abar_sum + agri_sum
                    perc_abar = (abar_sum / total_rec) if total_rec > 0 else 0
                    alert_status = "⚠️ ALERTA: >30%" if perc_abar > 0.30 else "OK"

                    ws_det.append([pdf_file.name, name_e, nit_e, nit_r, dte_val, m_name, alert_status])
                    new_count += 1

            progress_bar.progress((i + 1) / len(uploaded_pdfs))

        # 5. Write to Main Sheet securely
        for target_m_id, r_idx in row_map.items():
            data = batch_totals.get(target_m_id)
            if not data: continue

            if 'abar' in col_map and data['abar'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['abar'])
                target_cell.value = safe_float(target_cell.value) + data['abar']
            
            if 'agri' in col_map and data['agri'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['agri'])
                target_cell.value = safe_float(target_cell.value) + data['agri']

            if 'escuelas' in col_map and len(data['receptores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['escuelas'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['receptores'])
            
            if 'productores' in col_map and len(data['emisores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['productores'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['emisores'])

        # 6. Format "Extra Detalles" and "Items Sin Clasificar"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Format Extra Detalles
        for col in ws_det.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                cell.border = thin_border 
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_det.column_dimensions[col_letter].width = max_length + 2
        
        # Format Items Sin Clasificar
        for col in ws_unmatched.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                cell.border = thin_border 
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_unmatched.column_dimensions[col_letter].width = max_length + 2

        # 7. Final Export
        output = io.BytesIO()
        wb.save(output)
        
        # Count unmatched items (excluding header row)
        unmatched_count = ws_unmatched.max_row - 1 if ws_unmatched.max_row > 1 else 0
        
        success_msg = f"¡Proceso completado! {new_count} facturas procesadas y agregadas al Excel con éxito."
        if unmatched_count > 0:
            success_msg += f"""\n\n⚠️ {unmatched_count} items sin clasificar encontrados. Están en la tercera hoja del archivo de Excel, 'Items sin Clasificar', para revisión manual.
                            Los totales de esos productos no fueron agregados a la cantidad de la primera hoja"""
        
        st.success(success_msg)
        
        # Show warning for non-standard receipts that were skipped
        if skipped_non_standard:
            warning_msg = f"⚠️ **{len(skipped_non_standard)} factura(s) no estándar fueron ignoradas** (proformas, cotizaciones, u otros formatos no oficiales). Estas deben procesarse manualmente:\n\n"
            for pdf_name in skipped_non_standard:
                warning_msg += f"- {pdf_name}\n"
            st.warning(warning_msg)
        
        output.seek(0)
        st.download_button("Descargar Reporte Final", data=output.getvalue(), 
                           file_name="Reporte_MAGA_Actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Error crítico detectado: {e}")
